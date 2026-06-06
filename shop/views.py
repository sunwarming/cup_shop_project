import io
from django.shortcuts import render, get_object_or_404, redirect
from django.http import HttpResponse
from django.contrib.auth.decorators import login_required
from django.db.models import Q
from django.core.mail import EmailMessage

# Библиотеки для работы с генерацией чеков Excel
from openpyxl import Workbook
from openpyxl.styles import Font

from .models import Product, Category, Manufacturer, Cart, CartItem


# ==========================================
# 1. СТАТИЧЕСКИЕ / ИНФОРМАЦИОННЫЕ СТРАНИЦЫ
# ==========================================

def home_page(request):
    html_content = """
    <div style="font-family: Arial, sans-serif; margin: 40px;">
        <h1>Магазин тематических кружек и термосов</h1>
        <p>Добро пожаловать на главную страницу!</p>
        <hr>
        <h3>Быстрая навигация:</h3>
        <ul>
            <li><a href="/catalog/">🛒 Перейти в каталог товаров</a></li>
            <li><a href="/cart/">📦 Открыть мою корзину</a></li>
            <li><a href="/about-store/">ℹ️ О магазине</a></li>
            <li><a href="/about-author/">👨‍💻 Об авторе</a></li>
        </ul>
    </div>
    """
    return HttpResponse(html_content)


def about_store_page(request):
    text = "Лабораторная работа: Реализация CRUD-операций и интеграция со сторонними библиотеками в Django."
    return HttpResponse(text, content_type="text/plain; charset=utf-8")


def about_author_page(request):
    text = "Выполнил: Студент группы 88ТП Козик Николай"
    return HttpResponse(text, content_type="text/plain; charset=utf-8")


# ==========================================
# 2. КАТАЛОГ ТОВАРОВ И ФИЛЬТРАЦИЯ
# ==========================================

def product_list(request):
    query = request.GET.get('q', '')
    category_id = request.GET.get('category', '')
    
    products = Product.objects.all()
    categories = Category.objects.all()

    # Поиск по названию или описанию
    if query:
        products = products.filter(Q(name__icontains=query) | Q(description__icontains=query))
        
    # Фильтрация по категории
    if category_id:
        products = products.filter(category_id=category_id)

    return render(request, 'shop/product_list.html', {
        'products': products,
        'categories': categories,
        'query': query,
        'selected_category': category_id
    })


def product_detail(request, pk):
    product = get_object_or_404(Product, pk=pk)
    return render(request, 'shop/product_detail.html', {'product': product})


# ==========================================
# 3. ЛОГИКА РАБОТЫ КОРЗИНЫ (CART)
# ==========================================

@login_required(login_url='/admin/login/')
def cart_view(request):
    cart, created = Cart.objects.get_or_create(user=request.user)
    cart_items = cart.items.all()
    return render(request, 'shop/cart.html', {'cart': cart, 'cart_items': cart_items})


@login_required(login_url='/admin/login/')
def add_to_cart(request, product_id):
    product = get_object_or_404(Product, id=product_id)
    cart, created = Cart.objects.get_or_create(user=request.user)
    
    cart_item, item_created = CartItem.objects.get_or_create(cart=cart, product=product)
    
    if not item_created:
        if cart_item.quantity < product.stock:
            cart_item.quantity += 1
            cart_item.save()
    else:
        if product.stock >= 1:
            cart_item.quantity = 1
            cart_item.save()
        else:
            cart_item.delete()

    return redirect('cart_view')


@login_required(login_url='/admin/login/')
def update_cart(request, item_id):
    cart_item = get_object_or_404(CartItem, id=item_id, cart__user=request.user)
    if request.method == 'POST':
        new_quantity = int(request.POST.get('quantity', 1))
        if 1 <= new_quantity <= cart_item.product.stock:
            cart_item.quantity = new_quantity
            cart_item.save()
    return redirect('cart_view')


@login_required(login_url='/admin/login/')
def remove_from_cart(request, item_id):
    cart_item = get_object_or_404(CartItem, id=item_id, cart__user=request.user)
    cart_item.delete()
    return redirect('cart_view')


# ==========================================
# 4. ОФОРМЛЕНИЕ ЗАКАЗА, EXCEL И EMAIL
# ==========================================

@login_required(login_url='/admin/login/')
def checkout(request):
    cart = get_object_or_404(Cart, user=request.user)
    cart_items = cart.items.all()

    # Если корзина пуста, оформлять нечего
    if not cart_items:
        return redirect('product_list')

    if request.method == 'POST':
        address = request.POST.get('address', '')
        comment = request.POST.get('comment', '')

        # --- ГЕНЕРАЦИЯ EXCEL ЧЕКА (openpyxl) ---
        wb = Workbook()
        ws = wb.active
        ws.title = "Чек заказа"

        ws['A1'] = "ТОВАРНЫЙ ЧЕК"
        ws['A1'].font = Font(name='Arial', size=16, bold=True)
        ws['A2'] = f"Покупатель: {request.user.username} ({request.user.email})"
        ws['A3'] = f"Адрес доставки: {address}"
        
        headers = ["Товар", "Цена (BYN)", "Количество", "Стоимость"]
        ws.append([]) 
        ws.append(headers)
        
        for item in cart_items:
            ws.append([
                item.product.name,
                float(item.product.price),
                item.quantity,
                float(item.item_price)
            ])
            
            # Уменьшаем остатки на складе
            item.product.stock -= item.quantity
            item.product.save()

        ws.append([])
        ws.append(["ИТОГО К ОПЛАТЕ:", "", "", float(cart.total_price)])
        
        for col in ws.columns:
            max_len = max(len(str(cell.value or '')) for cell in col)
            col_letter = col[0].column_letter
            ws.column_dimensions[col_letter].width = max(max_len + 3, 12)

        excel_buffer = io.BytesIO()
        wb.save(excel_buffer)
        excel_buffer.seek(0)

        # --- ОТПРАВКА EMAIL ---
        subject = f"Заказ успешно оформлен! Чек для {request.user.username}"
        body = (
            f"Здравствуйте, {request.user.username}!\n\n"
            f"Ваш заказ успешно принят в обработку.\n"
            f"Адрес доставки: {address}\n"
            f"Комментарий: {comment}\n\n"
            f"Детализированный чек в формате Excel прикреплен к этому письму.\n"
            f"Спасибо за покупку в нашем магазине кружек и термосов!"
        )
        
        email = EmailMessage(
            subject=subject,
            body=body,
            from_email='shop@cupstore.by',
            to=[request.user.email if request.user.email else 'test@example.com'],
        )
        
        email.attach(
            filename=f"receipt_order_{request.user.id}.xlsx",
            content=excel_buffer.getvalue(),
            mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )
        
        email.send()

        # Очищаем корзину покупателя
        cart_items.delete()

        return HttpResponse("""
            <div style="text-align: center; margin-top: 50px; font-family: Arial;">
                <h1 style="color: green;">Заказ успешно оформлен!</h1>
                <p>Чек в формате Excel сгенерирован и отправлен вам на Email (отображен в консоли терминала).</p>
                <a href="/catalog/">Вернуться в каталог</a>
            </div>
        """)

    return render(request, 'shop/checkout.html', {'cart': cart, 'cart_items': cart_items})